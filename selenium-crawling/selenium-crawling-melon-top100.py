# 대기 및 sleep 처리를 위한 time 모듈
import time

import openpyxl
# Selenium 웹드라이버 모듈 임포트
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
# 크롬 웹드라이버 실행에 필요한 서비스 객체와 옵션 객체를 위한 모듈 임포트
from selenium.webdriver.chrome.service import Service
# 요소 선택(By) 및 키보드 조작(Keys)을 위한 모듈 임포트
from selenium.webdriver.common.by import By
# 웹드라이버 자동 설치 도우미 패키지 (webdriver-manager)
from webdriver_manager.chrome import ChromeDriverManager

# ChromeDriverManager를 이용하여 크롬 드라이버를 자동 설치하고 Service 객체로 초기화
customService = Service(ChromeDriverManager().install())

# 크롬 브라우저의 실행 옵션을 담는 객체 생성 (필요시 headless 등 설정 가능)
customOptions = Options()

# 위에서 정의한 서비스와 옵션을 사용하여 Chrome 브라우저 실행
browser = webdriver.Chrome(service=customService, options=customOptions)

URL = 'https://www.melon.com/'

browser.get(URL)

browser.implicitly_wait(10)

browser.find_element(By.XPATH,'//*[@id="conts"]/div[5]/div/ul/li[1]/div/div/div/a').click()
time.sleep(2)

xlsxFile = openpyxl.Workbook()
xlsxSheet = xlsxFile.active
xlsxSheet.cell(row=1, column=1).value = '제목'
xlsxSheet.cell(row=1, column=2).value = '아티스트'

for i in range(100):
    title = browser.find_element(By.XPATH,f'/html/body/div/div[3]/div/div/div[3]/form/div/table/tbody/tr[{i + 1}]/td[6]/div/div/div[1]/span/a').text
    artist = browser.find_element(By.XPATH,f'/html/body/div/div[3]/div/div/div[3]/form/div/table/tbody/tr[{i + 1}]/td[6]/div/div/div[2]/a').text

    print('i = ', i ,", 제목 = ",title, ', 아티스트 = ', artist)
    xlsxSheet.cell(row= i + 2, column=1).value = title
    xlsxSheet.cell(row= i + 2, column=2).value = artist

xlsxFile.save('melon-top100.xlsx')
