import openpyxl

# 엑셀파일 생성
xlsxFile = openpyxl.Workbook()

# 생성한 파일에서 시트 생성
xlsxSheet = xlsxFile.active
#
# # 시트 특정 셀에 데이터 입력
# xlsxSheet.cell(row=1, column=1).value = "hi"

# 반복문으로 데이터 입력
for i in range(10):
    xlsxSheet.cell(row= 1 , column= i + 1).value = i +1

# 저장
xlsxFile.save('result.xlsx')